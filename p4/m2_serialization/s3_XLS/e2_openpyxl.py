"""
Python Course - Part 4

https://github.com/egalli64/pyco

Module 2 - Serialization

Deserializing an XLSX file by openpyxl

https://openpyxl.readthedocs.io/en/stable/
"""

from openpyxl import load_workbook

PATH = "p4/m2_serialization/s3_XLS/"
FILENAME = PATH + "friends.xlsx"

book = load_workbook(FILENAME)
print(f"Worksheet names are {book.sheetnames}\n")

# Get the worksheet from the workbook
sheet = book[book.sheetnames[0]]

# Loop on each row in the worksheet
for row in sheet.rows:
    # Loop on each cell in the current row
    for cell in row:
        # Extract the value from the current cell
        print(cell.value, end=" ")
    print()
