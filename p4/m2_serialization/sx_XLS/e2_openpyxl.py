"""
Python Course - Part 4

https://github.com/egalli64/pyco

Module 2 - Serialization

Accessing XLSX file by openpyxl

https://openpyxl.readthedocs.io/en/stable/

pip install openpyxl
"""

from openpyxl import load_workbook

SOURCE = "p4/m2_serialization/sx_XLS/friends.xlsx"

book = load_workbook(SOURCE)
print("Worksheet names are", book.sheetnames)

print("Worksheet", book.sheetnames[0])
for row in book[book.sheetnames[0]].rows:
    for cell in row:
        print(cell.value, end=" ")
    print()
